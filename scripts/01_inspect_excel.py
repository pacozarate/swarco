#!/usr/bin/env python3
"""Inspect dbo_alhis.xlsx and write a traceable schema JSON."""

from __future__ import annotations

import hashlib
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel


ROOT = Path(__file__).resolve().parents[1]
RAW_FILE = ROOT / "data" / "raw" / "dbo_alhis.xlsx"
OUTPUT_FILE = ROOT / "data" / "processed" / "dbo_alhis_schema.json"
MAIN_SHEET = "dbo_alhis"
SAMPLE_ROWS = 20


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean_header(value: object) -> str:
    return "" if value is None else str(value).strip()


def json_value(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        serial = to_excel(value)
        return int(serial) if float(serial).is_integer() else serial
    return str(value)


def main() -> None:
    if not RAW_FILE.exists():
        raise FileNotFoundError(f"No existe el Excel original: {RAW_FILE}")

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    file_size = RAW_FILE.stat().st_size
    file_hash = sha256_file(RAW_FILE)

    workbook = load_workbook(RAW_FILE, read_only=True, data_only=True)
    sheets = []
    sample = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "columns": [],
        }
        if sheet_name == MAIN_SHEET:
            rows = sheet.iter_rows(values_only=True)
            headers = [clean_header(value) for value in next(rows, [])]
            sheet_info["columns"] = headers
            for row_index, row in enumerate(rows, start=1):
                if row_index > SAMPLE_ROWS:
                    break
                sample.append({headers[index] or f"column_{index + 1}": json_value(value) for index, value in enumerate(row)})
        sheets.append(sheet_info)

    workbook.close()

    schema = {
        "file": RAW_FILE.name,
        "path": str(RAW_FILE.relative_to(ROOT)),
        "size_bytes": file_size,
        "size_mb": round(file_size / (1024 * 1024), 2),
        "sha256": file_hash,
        "sheets": sheets,
        "main_sheet": MAIN_SHEET,
        "sample_first_20_rows": sample,
    }

    OUTPUT_FILE.write_text(json.dumps(schema, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Archivo: {RAW_FILE.name}")
    print(f"Tamaño: {file_size} bytes ({schema['size_mb']} MB)")
    print(f"SHA-256: {file_hash}")
    print("Hojas:")
    for sheet in sheets:
        print(f"  - {sheet['name']}: {sheet['max_row']} filas x {sheet['max_column']} columnas")
    main = next((sheet for sheet in sheets if sheet["name"] == MAIN_SHEET), None)
    if main:
        print(f"Cabeceras {MAIN_SHEET}: {', '.join(main['columns'])}")
        print(f"Filas aproximadas: {main['max_row']}")
        print(f"Columnas: {main['max_column']}")
    print(f"Schema JSON generado: {OUTPUT_FILE.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
