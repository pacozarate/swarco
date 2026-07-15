#!/usr/bin/env python3
"""Convert the dbo_alhis Excel sheet to Parquet in reproducible batches."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel


ROOT = Path(__file__).resolve().parents[1]
RAW_FILE = ROOT / "data" / "raw" / "dbo_alhis.xlsx"
OUTPUT_FILE = ROOT / "data" / "processed" / "dbo_alhis.parquet"
MAIN_SHEET = "dbo_alhis"
BATCH_SIZE = 25_000


def clean_header(value: object, index: int) -> str:
    text = "" if value is None else str(value).strip()
    return text or f"column_{index + 1}"


def preserve_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (datetime, date)):
        serial = to_excel(value)
        return str(int(serial)) if float(serial).is_integer() else str(serial)
    return str(value)


def batches(rows: Iterable[Sequence[object]], size: int) -> Iterable[list[Sequence[object]]]:
    batch: list[Sequence[object]] = []
    for row in rows:
        batch.append(row)
        if len(batch) >= size:
            yield batch
            batch = []
    if batch:
        yield batch


def main() -> None:
    if not RAW_FILE.exists():
        raise FileNotFoundError(f"No existe el Excel original: {RAW_FILE}")

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    if OUTPUT_FILE.exists():
        OUTPUT_FILE.unlink()

    workbook = load_workbook(RAW_FILE, read_only=True, data_only=True)
    if MAIN_SHEET not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"No existe la hoja requerida: {MAIN_SHEET}")

    sheet = workbook[MAIN_SHEET]
    row_iter = sheet.iter_rows(values_only=True)
    headers = [clean_header(value, index) for index, value in enumerate(next(row_iter))]
    schema = pa.schema([(header, pa.string()) for header in headers])

    writer = pq.ParquetWriter(
        OUTPUT_FILE,
        schema,
        compression="zstd",
        version="2.6",
        use_dictionary=True,
    )

    exported_rows = 0
    try:
        for batch_number, row_batch in enumerate(batches(row_iter, BATCH_SIZE), start=1):
            columns = {
                header: [preserve_value(row[index]) if index < len(row) else None for row in row_batch]
                for index, header in enumerate(headers)
            }
            table = pa.Table.from_pydict(columns, schema=schema)
            writer.write_table(table)
            exported_rows += table.num_rows
            print(f"Lote {batch_number}: {table.num_rows} filas exportadas; acumulado {exported_rows}")
    finally:
        writer.close()
        workbook.close()

    print(f"Parquet generado: {OUTPUT_FILE.relative_to(ROOT)}")
    print(f"Filas exportadas: {exported_rows}")
    print(f"Columnas exportadas: {len(headers)}")


if __name__ == "__main__":
    main()
