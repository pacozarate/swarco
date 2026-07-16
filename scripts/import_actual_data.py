from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


SOURCE_DIR = Path(
    "/Users/nuesopz/Documents/Descargas ICloud/PR23-0023_NG_LACROIX SERVICIO CALCULADORA COSTES 2023/"
    "03-CALCULO MECANICA/01-CALC_LACROIX/DATOS_A_IMPORTAR"
)
ROOT = Path(__file__).resolve().parents[1]
APP_DATA_CUTOFF = date(2020, 1, 1)
TRL_OVERRIDES = {
    "PN524AE01": {"widthMm": 1225, "heightMm": 780, "depthMm": 206.3},
    "PN531AE01": {"widthMm": 1030, "heightMm": 411, "depthMm": 155},
    "PN532AE01": {"widthMm": 1030, "heightMm": 411, "depthMm": 334},
    "PN533AE01": {"widthMm": 1200, "heightMm": 315, "depthMm": 125},
    "PN534AE01": {"widthMm": 1080, "heightMm": 290, "depthMm": 190},
}
TRL_CODE_OVERRIDES = {
    ("PN533A", "2L", "BANDEJA LEDS"): "PN533AE03",
}


def main() -> None:
    (ROOT / "data/trl").mkdir(parents=True, exist_ok=True)
    alart_rows_all = import_alart()
    gcesp_rows = import_gcesp(APP_DATA_CUTOFF)
    alhis_rows = import_alhis(APP_DATA_CUTOFF)
    ct_tft_rows = import_ct_tft()
    ct_led_rows = import_ct_led()
    trl_rows = import_trl()
    cplismat_rows = import_cplismat(alart_rows_all)
    needed_codes = collect_app_codes(cplismat_rows, gcesp_rows, alhis_rows, ct_tft_rows, ct_led_rows, trl_rows)
    alart_rows = [row for row in alart_rows_all if row["code"] in needed_codes]
    write_json(ROOT / "data/alart.json", alart_rows)
    write_json(ROOT / "data/alartdv.json", import_alartdv(needed_codes))
    write_json(ROOT / "data/cplismat.json", cplismat_rows)
    write_json(ROOT / "data/gcesp.json", gcesp_rows)
    write_json(ROOT / "data/alhis.json", alhis_rows)
    write_json(ROOT / "data/ct_tft.json", ct_tft_rows)
    write_json(ROOT / "data/ct_led.json", ct_led_rows)
    write_json(ROOT / "data/trl/pn-demo-trl.json", trl_rows)


def import_alart() -> list[dict[str, Any]]:
    rows = sheet_dicts(SOURCE_DIR / "Alart.xlsx", "alart")
    return [
        {
            "code": code(row.get("codart")),
            "description": text(row.get("descri1")),
            "pmp": number(row.get("pmedpon")),
            "type": text(row.get("tipart") or row.get("tigte")),
        }
        for row in rows
        if text(row.get("codart"))
    ]


def import_alartdv(allowed_codes: set[str] | None = None) -> list[dict[str, Any]]:
    rows = sheet_dicts(SOURCE_DIR / "Alart.xlsx", "dat.alfa.num")
    return [
        {
            "code": code(row.get("codart")),
            "dva17": value(row.get("dva17")),
            "dva18": value(row.get("dva18")),
            "dva19": value(row.get("dva19")),
            "dva20": number_or_none(row.get("dva20")),
            "dva37": value(row.get("dva37")),
            "dva38": value(row.get("dva38")),
            "dva39": value(row.get("dva39")),
            "dva40": value(row.get("dva40")),
        }
        for row in rows
        if text(row.get("codart")) and (allowed_codes is None or code(row.get("codart")) in allowed_codes)
    ]


def import_cplismat(alart_rows: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    rows = sheet_dicts(source_file("dbo_cplismat.xlsx", "DBO_CPLISMAT.xlsx"), "cplismat")
    tipart_by_code = {row["code"]: row.get("type", "") for row in (alart_rows or import_alart())}
    today = date.today()
    output = []
    for row in rows:
        parent = code(row.get("codsup"))
        child = code(row.get("codele"))
        quantity = row.get("cannec")
        valid_to = date_only(row.get("fecfin"))
        if not parent or not child or quantity is None or quantity == "" or valid_to is None or valid_to < today:
            continue
        output.append({
            "codsup": parent,
            "codele": child,
            "cannec": number(quantity),
            "fecfin": valid_to.isoformat(),
            "tipart": tipart_by_code.get(child, ""),
        })
    return output


def source_file(*names: str) -> Path:
    for name in names:
        path = SOURCE_DIR / name
        if path.exists():
            return path
    return SOURCE_DIR / names[0]


def date_only(item: Any) -> date | None:
    if isinstance(item, datetime):
        return item.date()
    if isinstance(item, date):
        return item
    parsed = date_value(item)
    if not parsed:
        return None
    try:
        return date.fromisoformat(parsed[:10])
    except ValueError:
        return None


def collect_app_codes(*datasets: list[dict[str, Any]]) -> set[str]:
    codes: set[str] = set()
    for rows in datasets:
        for row in rows:
            for key in ("code", "codsup", "codele", "root", "faCode", "dataCableCode", "powerCableCode"):
                current = code(row.get(key))
                if current:
                    codes.add(current)
    return codes


def import_gcesp(cutoff: date = APP_DATA_CUTOFF) -> list[dict[str, Any]]:
    rows = sheet_dicts(SOURCE_DIR / "dbo_gcesp.xlsx", "gcesp")
    today = date.today()
    output = [
        {
            "code": code(row.get("codart")),
            "description": text(row.get("des")),
            "price": number(row.get("pre")),
            "validFrom": date_value(row.get("fvdesde") or row.get("fecact")),
            "validTo": date_value(row.get("fvhasta")),
            "batch": value(row.get("lote")),
        }
        for row in rows
        if text(row.get("codart"))
        and (date_only(row.get("fvhasta")) or date.max) > today
        and (date_only(row.get("fvdesde") or row.get("fecact")) or date.min) < date(2027, 1, 1)
        and (date_only(row.get("fvhasta")) or date.max) >= cutoff
    ]
    return sorted(output, key=lambda row: row.get("validFrom") or "", reverse=True)


def import_alhis(cutoff: date = APP_DATA_CUTOFF) -> list[dict[str, Any]]:
    output = []
    for row in sheet_dicts(SOURCE_DIR / "dbo_alhis.xlsx", "dbo_alhis"):
        article_code = code(row.get("codart"))
        movement_date = date_only(row.get("fecmov"))
        if not article_code or text(row.get("moves")) != "E" or movement_date is None or movement_date < cutoff:
            continue
        output.append({
            "code": article_code,
            "date": movement_date.isoformat(),
            "quantity": number(row.get("cant")),
            "price": number(row.get("prec")),
            "movement": text(row.get("moves")),
            "supplier": value(row.get("clprfab")),
            "expiration": date_value(row.get("feccad")),
            "averageCost": number(row.get("premedpon")),
            "realCost": number(row.get("premedpon") or row.get("prec")),
        })
    return output


def import_ct_tft() -> list[dict[str, Any]]:
    rows = sheet_dicts(SOURCE_DIR / "Calculation Tool TFTS.xlsx", "CTOOLTFTS", header_row=2)
    return [
        {
            "code": code(row.get("Referencia SWARCO")),
            "description": " ".join(part for part in [text(row.get("Fabricante")), text(row.get("Referencia Fab"))] if part),
            "active": True,
            "manufacturer": text(row.get("Fabricante")),
            "inches": text(row.get("Tamaño")).replace("''", ""),
            "inchesNumber": number_or_none(clean_inches(row.get("Tamaño"))),
            "format": text(row.get("Aspect ratio")),
            "brightness": text(row.get("Luminosidad")),
            "resolution": normalized_resolution(row.get("Resolución")),
            "resolutionWidth": resolution_parts(row.get("Resolución"))[0],
            "resolutionHeight": resolution_parts(row.get("Resolución"))[1],
            "tempRange": text(row.get("Rango Temp.")),
            "visibleArea": normalize_size_text(row.get("Tamaño área activa"), row.get("Aspect ratio")),
            "outerSize": text(row.get("Tamaño_1")),
        }
        for row in rows
        if text(row.get("Referencia SWARCO"))
    ]


def import_ct_led() -> list[dict[str, Any]]:
    rows = sheet_dicts(SOURCE_DIR / "Calculation Tool LEDS.xlsx", "CTToolLEDS", header_row=2)
    return [
        {
            "code": code(row.get("Referencia SWARCO")),
            "description": text(row.get("col_3")),
            "color": text(row.get("Color de LED")),
            "moduleColumns": number(row.get("Resolución de módulo en x")),
            "moduleRows": number(row.get("Resolución de módulo en y")),
            "pitchX": number(row.get("Paso en x")),
            "pitchY": number(row.get("Paso en y")),
            "pasoxy": f"{number_text(row.get('Paso en x'))}|{number_text(row.get('Paso en y'))}",
            "currentModule": number(row.get("A total")),
            "faCode": code(row.get("Cod SWARCO FA")),
            "faVoltage": number(row.get("Vdc (V)")),
            "faCurrent": number(row.get("Adc (A)")),
            "dataCableCode": code(row.get("Cable datos (50mm)")),
            "powerCableCode": code(row.get("Conector alimentación")),
        }
        for row in rows
        if text(row.get("Referencia SWARCO"))
    ]


def import_trl() -> list[dict[str, Any]]:
    source = SOURCE_DIR / "Calculation Tool Model List - TRL.xlsx"
    rows = sheet_dicts(source, "Model List")
    output = []
    for row in rows:
        model = code(row.get("RAIZ"))
        group = value(row.get("GRUPO"))
        item_type = text(row.get("TIPO"))
        current_code = trl_code(model, group, item_type, code(row.get("CÓDIGO")))
        if not model or not current_code:
            continue
        item = {
            "model": model,
            "group": group,
            "groupLabel": text(row.get("GRUPO HERRAMIENT")),
            "type": item_type,
            "code": current_code,
            "description": text(row.get("DESCRIPCIÓN HERRAMIENTA") or row.get("DESCRIPCIÓN")),
            "longDescription": text(row.get("DESCRIPCIÓN")),
            "material": text(row.get("MATERIAL")),
            "glass": text(row.get("VIDRIO")),
            "widthMm": number(row.get("Largo Total")),
            "heightMm": number(row.get("Alto Total")),
            "depthMm": number(row.get("Fondo Total")),
            "image": text(row.get("IMAGE")),
            "root": current_code,
            "default": "1" if text(row.get("GRUPO")) == "0" else "",
        }
        item.update(TRL_OVERRIDES.get(current_code, {}))
        output.append(item)
    return output


def trl_code(model: str, group: Any, item_type: str, original_code: str) -> str:
    key = (model, str(group).upper(), item_type.upper())
    return TRL_CODE_OVERRIDES.get(key, original_code)


def sheet_dicts(path: Path, sheet_name: str, header_row: int = 1) -> Iterable[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = unique_headers([text(item) or f"col_{index + 1}" for index, item in enumerate(header_values)])
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
      if not any(cell not in (None, "") for cell in row):
          continue
      yield {headers[index]: cell for index, cell in enumerate(row[: len(headers)])}


def unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    output = []
    for header in headers:
        count = seen.get(header, 0)
        seen[header] = count + 1
        output.append(f"{header}_{count}" if count else header)
    return output


def write_json(path: Path, rows: list[dict[str, Any]]) -> None:
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2, default=serialize) + "\n", encoding="utf-8")
    print(f"{path.relative_to(ROOT)}: {len(rows)} rows")


def value(item: Any) -> Any:
    if isinstance(item, (datetime, date)):
        return item.isoformat()
    if item is None:
        return ""
    return item


def text(item: Any) -> str:
    if item is None:
        return ""
    if isinstance(item, float) and item.is_integer():
        return str(int(item))
    return str(item).strip()


def code(item: Any) -> str:
    return text(item).upper()


def number(item: Any) -> float:
    if item is None or item == "":
        return 0
    if isinstance(item, (int, float)):
        return item
    try:
        return float(str(item).replace(",", "."))
    except ValueError:
        return 0


def number_or_none(item: Any) -> float | None:
    if item is None or item == "":
        return None
    if isinstance(item, (int, float)):
        return item
    current = str(item).strip()
    if "." in current and "," not in current:
        current = current.replace(".", ",")
    try:
        return float(current.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def clean_inches(item: Any) -> str:
    current = text(item)
    for token in ("''", '"', "'", " in", "in", " "):
        current = current.replace(token, "")
    return current


def resolution_parts(item: Any) -> tuple[float | None, float | None]:
    current = text(item).upper().replace("×", "X").replace("*", "X").replace("x", "X").replace(" ", "")
    if "X" not in current:
        return (None, None)
    left, right = current.split("X", 1)
    return (number_or_none(left), number_or_none(right))


def normalized_resolution(item: Any) -> str:
    width, height = resolution_parts(item)
    if width is None or height is None:
        return text(item)
    return f"{format_size_number(width)}X{format_size_number(height)}"


def number_text(item: Any) -> str:
    current = number(item)
    if float(current).is_integer():
        return str(int(current))
    return str(current).replace(".", ",")


def normalize_size_text(size: Any, aspect_ratio: Any = "") -> str:
    current = text(size)
    if "x" not in current.lower():
        return current
    left, right = current.lower().split("x", 1)
    width = number(left)
    height = number(right)
    ratio = aspect_ratio_value(aspect_ratio)
    if not width or not height or not ratio:
        return current
    measured = width / height
    if measured > ratio * 2.5:
        width = height * ratio
    elif measured < ratio / 2.5:
        height = width / ratio
    return f"{format_size_number(width)}x{format_size_number(height)}"


def aspect_ratio_value(item: Any) -> float:
    parts = text(item).replace(",", ".").split(":")
    if len(parts) != 2:
        return 0
    left, right = number(parts[0]), number(parts[1])
    return left / right if right else 0


def format_size_number(item: float) -> str:
    return str(int(item))


def date_value(item: Any) -> str:
    if isinstance(item, (datetime, date)):
        return item.date().isoformat() if isinstance(item, datetime) else item.isoformat()
    return text(item)


def comparable_date(item: Any) -> str:
    return date_value(item) or "0001-01-01"


def serialize(item: Any) -> Any:
    if isinstance(item, (datetime, date)):
        return date_value(item)
    return item


if __name__ == "__main__":
    main()
